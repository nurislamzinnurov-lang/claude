import re
import time
import random
import threading
import queue
import json
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urljoin, parse_qs, unquote

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7; rv:120.0) Gecko/20100101 Firefox/120.0",
]

DEFAULT_CATEGORIES = [
    ("Автодилер BMW",            "официальный дилер BMW {city}"),
    ("Автодилер Mercedes",       "официальный дилер Mercedes {city}"),
    ("Автодилер Lexus",          "официальный дилер Lexus {city}"),
    ("Автодилер Audi",           "официальный дилер Audi {city}"),
    ("Автодилер Porsche",        "официальный дилер Porsche {city}"),
    ("Автосалон премиум",        "премиум автосалон {city}"),
    ("Элитная недвижимость",     "агентство элитной недвижимости {city}"),
    ("Застройщик премиум",       "застройщик премиум класса {city}"),
    ("Ювелирный бутик",          "ювелирный бутик {city}"),
    ("Швейцарские часы",         "швейцарские часы {city} бутик"),
    ("Private banking",          "private banking {city}"),
    ("Премиум банк",             "премиум банк {city}"),
    ("Премиум клиника",          "частная клиника премиум {city}"),
    ("Стоматология платная",     "платная стоматология {city}"),
    ("Дизайн интерьера",         "студия дизайна интерьера премиум {city}"),
    ("Бутик-отель",              "бутик отель {city}"),
    ("Отель 5*",                 "отель 5 звезд {city}"),
    ("Бизнес-школа",             "бизнес школа {city}"),
    ("Event-агентство",          "организация мероприятий {city}"),
    ("Премиум e-commerce",       "интернет-магазин премиум {city}"),
]

EXCLUDED_DOMAINS = {
    "2gis", "google", "youtube", "vk.com", "vk.ru", "facebook", "instagram",
    "twitter", "x.com", "ok.ru", "yandex", "ya.ru", "telegram", "t.me",
    "wikipedia", "tripadvisor", "booking", "avito", "cian", "yell", "zoon",
    "flamp", "spr.ru", "rutube", "dzen", "pikabu", "linkedin", "pinterest",
    "tiktok", "wb.ru", "wildberries", "ozon", "drom", "auto.ru", "irr.ru",
    "youla", "domclick", "n1.ru", "bing.com", "duckduckgo", "msn.com",
    "yapokupayu", "vc.ru", "spravker", "jsprav", "edem-v-gosti",
    "trip.com", "tutu.ru", "101hotels", "otello", "alean", "ostrovok",
    "travel.yandex", "vsedoma", "blog.", "habr.com", "pikabu",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
IMG_EXT = ("png", "jpg", "jpeg", "gif", "svg", "webp", "ico", "bmp", "tiff", "css", "js")
EMAIL_PRIORITY = ("marketing@", "info@", "sales@", "pr@", "director@",
                  "manager@", "office@", "contact@", "hello@", "mail@", "reception@")
JUNK_EMAILS = ("example.", "sentry", "wixpress", "yourdomain", "noreply", "no-reply",
               "donotreply", "@2x", "@3x", "domain.com", "site.com", "test@",
               "u003e", "u003c", "react.", ".png", ".jpg", ".jpeg", ".gif", ".svg")
CONTACT_KEYWORDS = ("контакт", "contact", "о нас", "about", "связ", "написать",
                    "наши контакты")
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".premium_search.json")


def random_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }


def ok_host(url):
    try:
        host = urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return False
    if not host or "." not in host:
        return False
    for bad in EXCLUDED_DOMAINS:
        if bad in host:
            return False
    return True


def unwrap_ddg(href):
    if not href:
        return ""
    if href.startswith("//duckduckgo.com/l/?") or href.startswith("//html.duckduckgo.com/l/?"):
        href = "https:" + href
    if "duckduckgo.com/l/" in href:
        qs = parse_qs(urlparse(href).query)
        if "uddg" in qs:
            return unquote(qs["uddg"][0])
    return href


def ddg_lite_search(query, max_results, log):
    sess = requests.Session()
    sess.headers.update(random_headers())
    sess.headers["Referer"] = "https://duckduckgo.com/"
    out = []
    seen_hosts = set()
    for page in range((max_results // 10) + 2):
        try:
            data = {"q": query, "kl": "ru-ru"}
            if page > 0:
                data["s"] = str(page * 30)
                data["dc"] = str(page * 30 + 1)
            r = sess.post("https://lite.duckduckgo.com/lite/", data=data, timeout=20)
        except Exception as e:
            log(f"  lite-ddg error: {e}")
            break
        if r.status_code != 200:
            break
        soup = BeautifulSoup(r.text, "html.parser")
        page_count = 0
        for a in soup.find_all("a", href=True):
            url = unwrap_ddg(a["href"])
            if not url.startswith("http"):
                continue
            if "duckduckgo.com" in url:
                continue
            if not ok_host(url):
                continue
            host = urlparse(url).netloc.lower().replace("www.", "")
            if host in seen_hosts:
                continue
            seen_hosts.add(host)
            root = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
            out.append((root, a.get_text(strip=True) or host))
            page_count += 1
            if len(out) >= max_results:
                return out
        if page_count == 0:
            break
        time.sleep(random.uniform(1.0, 2.0))
    return out


def ddg_html_search(query, max_results, log):
    sess = requests.Session()
    sess.headers.update(random_headers())
    out = []
    seen_hosts = set()
    try:
        r = sess.post("https://html.duckduckgo.com/html/",
                      data={"q": query, "kl": "ru-ru"}, timeout=20)
    except Exception as e:
        log(f"  html-ddg error: {e}")
        return out
    if r.status_code != 200:
        return out
    soup = BeautifulSoup(r.text, "html.parser")
    for a in soup.select("a.result__a"):
        url = unwrap_ddg(a.get("href", ""))
        if not url.startswith("http") or not ok_host(url):
            continue
        host = urlparse(url).netloc.lower().replace("www.", "")
        if host in seen_hosts:
            continue
        seen_hosts.add(host)
        root = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
        out.append((root, a.get_text(strip=True) or host))
        if len(out) >= max_results:
            break
    return out


def search(query, max_results, log):
    out = ddg_lite_search(query, max_results, log)
    if len(out) < 3:
        log(f"  lite вернул {len(out)}, пробую html-ddg...")
        time.sleep(random.uniform(1.5, 3.0))
        more = ddg_html_search(query, max_results, log)
        seen = {urlparse(u).netloc.lower().replace("www.", "") for u, _ in out}
        for u, n in more:
            host = urlparse(u).netloc.lower().replace("www.", "")
            if host not in seen:
                out.append((u, n))
                seen.add(host)
    return out[:max_results]


def filter_emails(emails):
    cleaned = []
    seen = set()
    for e in emails:
        e = e.strip().strip(".,;:")
        low = e.lower()
        if low in seen:
            continue
        if any(low.endswith("." + ext) for ext in IMG_EXT):
            continue
        if any(j in low for j in JUNK_EMAILS):
            continue
        if len(e) > 60 or len(e) < 6:
            continue
        seen.add(low)
        cleaned.append(e)

    def rank(e):
        low = e.lower()
        for i, p in enumerate(EMAIL_PRIORITY):
            if low.startswith(p):
                return i
        return len(EMAIL_PRIORITY)

    cleaned.sort(key=rank)
    return cleaned


def fetch(url, timeout=12):
    try:
        r = requests.get(url, headers=random_headers(), timeout=timeout,
                         allow_redirects=True, verify=True)
        if r.status_code == 200 and r.text:
            return r.text, r.url
    except requests.exceptions.SSLError:
        try:
            r = requests.get(url, headers=random_headers(), timeout=timeout,
                             allow_redirects=True, verify=False)
            if r.status_code == 200 and r.text:
                return r.text, r.url
        except Exception:
            pass
    except Exception:
        pass
    return None, None


def find_contact_links(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    links = []
    seen = set()
    base_host = urlparse(base_url).netloc
    for a in soup.find_all("a", href=True):
        text = (a.get_text() or "").lower()
        href = a["href"].lower()
        if any(k in text for k in CONTACT_KEYWORDS) or any(k in href for k in CONTACT_KEYWORDS):
            try:
                full = urljoin(base_url, a["href"])
            except Exception:
                continue
            if urlparse(full).netloc != base_host:
                continue
            if full in seen or full == base_url:
                continue
            seen.add(full)
            links.append(full)
        if len(links) >= 3:
            break
    return links


def extract_title(html):
    try:
        soup = BeautifulSoup(html, "html.parser")
        if soup.title and soup.title.string:
            return re.sub(r"\s+", " ", soup.title.string).strip()[:150]
    except Exception:
        pass
    return ""


def scrape_site(url, category, fallback_name, stop_event):
    if stop_event.is_set():
        return None
    html, final_url = fetch(url)
    if not html:
        return {"name": fallback_name, "category": category, "email": "",
                "site": url, "all_emails": ""}

    name = extract_title(html) or fallback_name
    emails = EMAIL_RE.findall(html)

    if not any(e.lower().startswith(EMAIL_PRIORITY) for e in emails):
        for link in find_contact_links(html, final_url or url):
            if stop_event.is_set():
                break
            chtml, _ = fetch(link)
            if chtml:
                emails.extend(EMAIL_RE.findall(chtml))
                if any(e.lower().startswith(EMAIL_PRIORITY) for e in emails):
                    break
            time.sleep(random.uniform(0.2, 0.5))

    emails = filter_emails(emails)
    best = emails[0] if emails else ""
    all_str = "; ".join(emails[:5])
    return {"name": name, "category": category, "email": best, "site": url,
            "all_emails": all_str}


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Поиск премиум-бизнесов РФ — email парсер")
        self.root.geometry("1180x760")
        self.root.minsize(960, 640)

        self.results = []
        self.stop_event = threading.Event()
        self.worker_thread = None
        self.ui_queue = queue.Queue()

        self.city_var = tk.StringVar(value="Санкт-Петербург")
        self.results_var = tk.IntVar(value=15)
        self.threads_var = tk.IntVar(value=12)
        self.delay_var = tk.DoubleVar(value=2.0)
        self.cat_vars = {}

        self._build_ui()
        self._load_config()
        self.root.after(100, self._drain_queue)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_ui(self):
        try:
            style = ttk.Style()
            if "clam" in style.theme_names():
                style.theme_use("clam")
        except Exception:
            pass

        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=tk.X)
        ttk.Label(top, text="Город:").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.city_var, width=28).grid(row=0, column=1, padx=(6, 16), sticky="w")

        ttk.Label(top, text="Результатов на запрос:").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(top, from_=5, to=50, increment=5, textvariable=self.results_var, width=6).grid(row=0, column=3, padx=(6, 16))

        ttk.Label(top, text="Потоков:").grid(row=0, column=4, sticky="w")
        ttk.Spinbox(top, from_=1, to=30, textvariable=self.threads_var, width=6).grid(row=0, column=5, padx=(6, 16))

        ttk.Label(top, text="Пауза между запросами (сек):").grid(row=0, column=6, sticky="w")
        ttk.Spinbox(top, from_=0.5, to=10.0, increment=0.5, textvariable=self.delay_var, width=6, format="%.1f").grid(row=0, column=7, padx=(6, 0))

        body = ttk.Frame(self.root, padding=(10, 0, 10, 0))
        body.pack(fill=tk.BOTH, expand=True)

        left = ttk.LabelFrame(body, text="Категории поиска", padding=8)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))

        btn_row = ttk.Frame(left)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="Все", width=8, command=lambda: self._toggle_all(True)).pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(btn_row, text="Снять", width=8, command=lambda: self._toggle_all(False)).pack(side=tk.LEFT)

        canvas_box = ttk.Frame(left)
        canvas_box.pack(fill=tk.BOTH, expand=True, pady=(6, 0))
        cat_canvas = tk.Canvas(canvas_box, width=260, highlightthickness=0)
        cat_scroll = ttk.Scrollbar(canvas_box, orient=tk.VERTICAL, command=cat_canvas.yview)
        cat_canvas.configure(yscrollcommand=cat_scroll.set)
        cat_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cat_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        cat_inner = ttk.Frame(cat_canvas)
        cat_canvas.create_window((0, 0), window=cat_inner, anchor="nw")
        cat_inner.bind("<Configure>", lambda e: cat_canvas.configure(scrollregion=cat_canvas.bbox("all")))

        for label, _ in DEFAULT_CATEGORIES:
            v = tk.BooleanVar(value=True)
            self.cat_vars[label] = v
            ttk.Checkbutton(cat_inner, text=label, variable=v).pack(anchor="w", pady=1)

        ttk.Separator(left, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)
        ttk.Label(left, text="Свои запросы (по одному на строку, {city} = город):").pack(anchor="w")
        self.custom_text = scrolledtext.ScrolledText(left, width=34, height=6, wrap=tk.WORD, font=("Consolas", 9))
        self.custom_text.pack(fill=tk.X, pady=(4, 0))
        self.custom_text.insert("1.0", "")

        right = ttk.Frame(body)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        actions = ttk.Frame(right)
        actions.pack(fill=tk.X, pady=(0, 6))
        self.start_btn = ttk.Button(actions, text="▶  Запустить поиск", command=self.start_search)
        self.start_btn.pack(side=tk.LEFT)
        self.stop_btn = ttk.Button(actions, text="■  Стоп", command=self.stop_search, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=6)
        self.save_btn = ttk.Button(actions, text="💾  Сохранить Excel", command=self.save_excel, state=tk.DISABLED)
        self.save_btn.pack(side=tk.LEFT, padx=6)
        ttk.Button(actions, text="Очистить", command=self.clear_results).pack(side=tk.LEFT, padx=6)
        self.count_var = tk.StringVar(value="")
        ttk.Label(actions, textvariable=self.count_var, foreground="#555").pack(side=tk.RIGHT)

        cols = ("name", "category", "email", "site")
        tree_frame = ttk.Frame(right)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=14)
        for col, title, width in zip(
            cols,
            ("Название", "Категория", "Email", "Сайт"),
            (320, 170, 220, 280),
        ):
            self.tree.heading(col, text=title, command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=width, anchor=tk.W)
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        self.tree.tag_configure("has_email", background="#e9fbe9")
        self.tree.tag_configure("no_email", background="#ffffff")
        self.tree.bind("<Double-1>", self._open_site)

        log_frame = ttk.LabelFrame(right, text="Лог", padding=4)
        log_frame.pack(fill=tk.BOTH, expand=False, pady=(8, 0))
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, wrap=tk.WORD,
                                                  font=("Consolas", 9), state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        bottom = ttk.Frame(self.root, padding=(10, 4, 10, 8))
        bottom.pack(fill=tk.X)
        self.status_var = tk.StringVar(value="Готов к работе. Выберите категории и нажмите «Запустить поиск».")
        ttk.Label(bottom, textvariable=self.status_var).pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(bottom, mode="determinate", length=320)
        self.progress.pack(side=tk.RIGHT)

    def _toggle_all(self, state):
        for v in self.cat_vars.values():
            v.set(state)

    def _drain_queue(self):
        try:
            while True:
                fn = self.ui_queue.get_nowait()
                try: fn()
                except Exception: pass
        except queue.Empty:
            pass
        self.root.after(100, self._drain_queue)

    def _post(self, fn):
        self.ui_queue.put(fn)

    def set_status(self, text):
        self._post(lambda: self.status_var.set(text))

    def set_progress(self, value, maximum=None):
        def upd():
            if maximum is not None:
                self.progress["maximum"] = maximum
            self.progress["value"] = value
        self._post(upd)

    def log(self, msg):
        ts = time.strftime("%H:%M:%S")
        def upd():
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"[{ts}] {msg}\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state=tk.DISABLED)
        self._post(upd)

    def add_row(self, item):
        def upd():
            tag = "has_email" if item["email"] else "no_email"
            self.tree.insert("", tk.END, values=(item["name"], item["category"],
                                                  item["email"], item["site"]), tags=(tag,))
            with_email = sum(1 for r in self.results if r["email"])
            self.count_var.set(f"Найдено: {len(self.results)}  •  с email: {with_email}")
        self._post(upd)

    def clear_results(self):
        self.results.clear()
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.count_var.set("")
        self.save_btn.config(state=tk.DISABLED)

    def _sort_by(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        items.sort(key=lambda x: x[0].lower())
        for i, (_, k) in enumerate(items):
            self.tree.move(k, "", i)

    def _open_site(self, _evt):
        sel = self.tree.selection()
        if not sel:
            return
        url = self.tree.item(sel[0], "values")[3]
        if url:
            import webbrowser
            webbrowser.open(url)

    def _build_queries(self):
        city = self.city_var.get().strip() or "Санкт-Петербург"
        out = []
        for label, tmpl in DEFAULT_CATEGORIES:
            if self.cat_vars.get(label) and self.cat_vars[label].get():
                out.append((tmpl.format(city=city), label))
        for line in self.custom_text.get("1.0", tk.END).splitlines():
            line = line.strip()
            if not line:
                continue
            q = line.replace("{city}", city).replace("{город}", city)
            out.append((q, "Свой запрос"))
        return out

    def start_search(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return
        queries = self._build_queries()
        if not queries:
            messagebox.showwarning("Нет запросов", "Выберите хотя бы одну категорию или добавьте свой запрос.")
            return
        self.clear_results()
        self.stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.save_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self._save_config()
        self.log(f"Старт. Запросов: {len(queries)}, потоков: {self.threads_var.get()}, "
                 f"результатов/запрос: {self.results_var.get()}.")
        self.worker_thread = threading.Thread(target=self._run, args=(queries,), daemon=True)
        self.worker_thread.start()

    def stop_search(self):
        self.stop_event.set()
        self.set_status("Останавливаюсь, дождитесь окончания текущих запросов...")
        self.log("Получен сигнал остановки.")

    def _run(self, queries):
        max_results = max(5, int(self.results_var.get()))
        threads = max(1, min(30, int(self.threads_var.get())))
        delay = max(0.5, float(self.delay_var.get()))

        self.set_status(f"Сбор ссылок: 0/{len(queries)}")
        self.set_progress(0, len(queries))
        all_targets = []
        for idx, (query, category) in enumerate(queries, 1):
            if self.stop_event.is_set():
                break
            self.set_status(f"Поиск {idx}/{len(queries)}: {query}")
            self.log(f"DDG ← {query}")
            try:
                urls = search(query, max_results, self.log)
            except Exception as e:
                self.log(f"  ошибка: {e}")
                urls = []
            self.log(f"  → найдено сайтов: {len(urls)}")
            for url, name in urls:
                all_targets.append((url, category, name))
            self.set_progress(idx)
            if idx < len(queries):
                time.sleep(delay + random.uniform(0, 0.6))

        seen, unique_targets = set(), []
        for url, cat, name in all_targets:
            key = urlparse(url).netloc.lower().replace("www.", "")
            if not key or key in seen:
                continue
            seen.add(key)
            unique_targets.append((url, cat, name))

        total = len(unique_targets)
        self.log(f"Уникальных сайтов: {total}. Извлекаю email в {threads} потоков.")
        self.set_status(f"Сканирование сайтов: 0/{total}")
        self.set_progress(0, max(total, 1))

        if not total:
            self.log("Ничего не найдено. Проверьте интернет/VPN.")
            self._finish()
            return

        with ThreadPoolExecutor(max_workers=threads) as ex:
            futures = {ex.submit(scrape_site, url, cat, name, self.stop_event): url
                       for url, cat, name in unique_targets}
            done = 0
            for fut in as_completed(futures):
                if self.stop_event.is_set():
                    for f in futures:
                        f.cancel()
                    break
                try:
                    item = fut.result()
                except Exception as e:
                    item = None
                    self.log(f"  scrape error: {e}")
                done += 1
                self.set_progress(done)
                if item:
                    self.results.append(item)
                    self.add_row(item)
                    if item["email"]:
                        self.log(f"  ✓ {item['email']}  ←  {item['site']}")
                with_email = sum(1 for r in self.results if r["email"])
                self.set_status(f"Сканирование: {done}/{total}  •  с email: {with_email}")

        self._finish()

    def _finish(self):
        self.results.sort(key=lambda r: (0 if r["email"] else 1, r["category"], r["name"]))

        def upd():
            for i in self.tree.get_children():
                self.tree.delete(i)
            for r in self.results:
                tag = "has_email" if r["email"] else "no_email"
                self.tree.insert("", tk.END, values=(r["name"], r["category"],
                                                     r["email"], r["site"]), tags=(tag,))
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.save_btn.config(state=tk.NORMAL if self.results else tk.DISABLED)
            with_email = sum(1 for r in self.results if r["email"])
            self.status_var.set(f"Готово. Всего: {len(self.results)} • с email: {with_email}")
            self.count_var.set(f"Найдено: {len(self.results)}  •  с email: {with_email}")
        self._post(upd)
        self.log(f"Готово. Всего сайтов: {len(self.results)}, с email: "
                 f"{sum(1 for r in self.results if r['email'])}.")

    def save_excel(self):
        if not self.results:
            messagebox.showinfo("Нет данных", "Сначала выполните поиск.")
            return
        city = self.city_var.get().strip().replace(" ", "_") or "city"
        default = f"premium_{city}_{time.strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default,
        )
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Бизнесы"
            headers = ["Название", "Категория", "Email", "Сайт", "Все email"]
            ws.append(headers)
            head_font = Font(bold=True, color="FFFFFF")
            head_fill = PatternFill("solid", fgColor="2A5298")
            for c in ws[1]:
                c.font = head_font
                c.fill = head_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
            good_fill = PatternFill("solid", fgColor="E9FBE9")
            for r in self.results:
                row = [r["name"], r["category"], r["email"], r["site"], r.get("all_emails", "")]
                ws.append(row)
                if r["email"]:
                    for c in ws[ws.max_row]:
                        c.fill = good_fill
            for col, width in zip("ABCDE", (48, 24, 32, 42, 50)):
                ws.column_dimensions[col].width = width
            ws.freeze_panes = "A2"
            wb.save(path)
            messagebox.showinfo("Сохранено", f"Файл сохранён:\n{path}")
            self.log(f"Файл сохранён: {path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _save_config(self):
        try:
            data = {
                "city": self.city_var.get(),
                "results": self.results_var.get(),
                "threads": self.threads_var.get(),
                "delay": self.delay_var.get(),
                "categories": {k: v.get() for k, v in self.cat_vars.items()},
                "custom": self.custom_text.get("1.0", tk.END).strip(),
            }
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.city_var.set(data.get("city", self.city_var.get()))
            self.results_var.set(int(data.get("results", self.results_var.get())))
            self.threads_var.set(int(data.get("threads", self.threads_var.get())))
            self.delay_var.set(float(data.get("delay", self.delay_var.get())))
            cats = data.get("categories", {})
            for k, v in self.cat_vars.items():
                if k in cats:
                    v.set(bool(cats[k]))
            custom = data.get("custom", "")
            if custom:
                self.custom_text.delete("1.0", tk.END)
                self.custom_text.insert("1.0", custom)
        except Exception:
            pass

    def _on_close(self):
        self._save_config()
        self.stop_event.set()
        self.root.destroy()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
